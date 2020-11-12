import openpyxl
import json

IPAM = openpyxl.load_workbook("IPAMExport/Structure.xlsx")
IPsheet = IPAM["Subnets"]

#https://github.com/jbrooksuk/JSON-Airports , json file for airport json database

with open('airports.json') as f:
    airports = json.load(f)

for i in range(1,IPsheet.max_row):
    
   IATA = str(IPsheet.cell(row=i,column=1).value)
   
   for x in airports:
    
       try:
        if IATA in x['iata']:
          print(IATA+" is "+x['name'])
          IPsheet.cell(row=i,column=2).value = x['name']
          print("Wrote to Sheet")
       except:
          print("IATA "+IATA+" not found")

IPAM.save("IPAMExport/Structure.xlsx")
print("Saved Book")
